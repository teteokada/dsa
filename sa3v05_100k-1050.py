# -*- coding: utf-8 -*- 

###########################################################################
## Score Analyzer 3_v05 @gses.jp
###########################################################################

import wx
import wx.xrc
import wx.richtext
import wx.grid
import os
import pandas as pd
import xlrd
import xlwt
import numpy as np
from decimal import Decimal
from matplotlib.figure import Figure
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigCanvas
import math
from openpyxl import Workbook
#from openpyxl.compat import range
#from openpyxl.cell import get_column_letter

import logging

###########################################################################
## Class MyFrame1
###########################################################################

class MyFrame1 ( wx.Frame ):

	def __init__( self, parent ):
		maxccs = 100000
		maxchains = 1050

		wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Score Analyzer 3_v05", pos = wx.DefaultPosition, size = wx.Size( 1247,820 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

		self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

		gbSizer2 = wx.GridBagSizer( 0, 0 )
		gbSizer2.SetFlexibleDirection( wx.BOTH )
		gbSizer2.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )

		sbSizer15 = wx.StaticBoxSizer( wx.StaticBox( self, wx.ID_ANY, u"Inputs" ), wx.VERTICAL )

		bSizer5 = wx.BoxSizer( wx.HORIZONTAL )


		bSizer5.Add( ( 10, 0), 0, wx.EXPAND, 5 )

		self.m_textCtrl3 = wx.TextCtrl( sbSizer15.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 180,-1 ), wx.TE_CENTER )
		bSizer5.Add( self.m_textCtrl3, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 5 )


		bSizer5.Add( ( 10, 0), 1, 0, 5 )

		self.m_button1 = wx.Button( sbSizer15.GetStaticBox(), wx.ID_ANY, u"Load PDB", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		bSizer5.Add( self.m_button1, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 5 )

		self.m_button16 = wx.Button( sbSizer15.GetStaticBox(), wx.ID_ANY, u"Load CIF", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		bSizer5.Add( self.m_button16, 0, wx.ALL, 5 )
        
		self.m_button81 = wx.Button( sbSizer15.GetStaticBox(), wx.ID_ANY, u"All Clear", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		bSizer5.Add( self.m_button81, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 5 )


		sbSizer15.Add( bSizer5, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.EXPAND, 5 )

		self.m_richText2 = wx.richtext.RichTextCtrl( sbSizer15.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 550,292 ), 0|wx.BORDER_SUNKEN|wx.HSCROLL|wx.VSCROLL|wx.WANTS_CHARS )

		sbSizer15.Add( self.m_richText2, 1, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 5 )

		gSizer1 = wx.GridSizer( 2, 5, 0, 0 )

		self.m_staticText8 = wx.StaticText( sbSizer15.GetStaticBox(), wx.ID_ANY, u"Num. Chains", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticText8.Wrap( -1 )

		gSizer1.Add( self.m_staticText8, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText9 = wx.StaticText( sbSizer15.GetStaticBox(), wx.ID_ANY, u"Num. CA", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticText9.Wrap( -1 )

		gSizer1.Add( self.m_staticText9, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText6 = wx.StaticText( sbSizer15.GetStaticBox(), wx.ID_ANY, u"Num. C-C", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticText6.Wrap( -1 )

		gSizer1.Add( self.m_staticText6, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText61 = wx.StaticText( sbSizer15.GetStaticBox(), wx.ID_ANY, u"1st posi.", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticText61.Wrap( -1 )

		gSizer1.Add( self.m_staticText61, 0, wx.ALIGN_CENTER|wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_TOP|wx.ALL, 0 )

		self.m_staticText7 = wx.StaticText( sbSizer15.GetStaticBox(), wx.ID_ANY, u"Chain ID", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticText7.Wrap( -1 )

		gSizer1.Add( self.m_staticText7, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 0 )

		self.m_textCtrl31 = wx.TextCtrl( sbSizer15.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 40,-1 ), wx.TE_CENTER )
		gSizer1.Add( self.m_textCtrl31, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 0 )

		self.m_textCtrl6 = wx.TextCtrl( sbSizer15.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), wx.TE_CENTER )
		self.m_textCtrl6.SetMinSize( wx.Size( 60,-1 ) )

		gSizer1.Add( self.m_textCtrl6, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 0 )

		self.m_textCtrl2 = wx.TextCtrl( sbSizer15.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 80,-1 ), wx.TE_CENTER )
		gSizer1.Add( self.m_textCtrl2, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 0 )

		self.m_textCtrl7 = wx.TextCtrl( sbSizer15.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), wx.TE_CENTER )
		gSizer1.Add( self.m_textCtrl7, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 0 )

		self.m_textCtrl4 = wx.TextCtrl( sbSizer15.GetStaticBox(), wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 40,-1 ), wx.TE_CENTER )
		gSizer1.Add( self.m_textCtrl4, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 0 )


		sbSizer15.Add( gSizer1, 0, wx.EXPAND, 5 )


		gbSizer2.Add( sbSizer15, wx.GBPosition( 0, 0 ), wx.GBSpan( 1, 1 ), wx.EXPAND|wx.LEFT|wx.RIGHT|wx.TOP, 5 )

		sbSizer16 = wx.StaticBoxSizer( wx.StaticBox( self, wx.ID_ANY, u"Distances" ), wx.VERTICAL )

		self.m_grid2 = wx.grid.Grid( sbSizer16.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition, wx.Size( 613,655 ), 0 )

		# Grid
		self.m_grid2.CreateGrid( maxccs, maxchains+5 )
		self.m_grid2.EnableEditing( True )
		self.m_grid2.EnableGridLines( False )
		self.m_grid2.EnableDragGridSize( False )
		self.m_grid2.SetMargins( 0, 0 )

		# Columns
		# Columns
		self.m_grid2.SetColSize( 3, 55 )
		self.m_grid2.SetColSize( 4, 55 )

		self.m_grid2.EnableDragColMove( False )
		self.m_grid2.EnableDragColSize( True )
		self.m_grid2.SetColLabelSize( 25 )
		self.m_grid2.SetColLabelValue( 0, u"X" )
		self.m_grid2.SetColLabelValue( 1, u"Y" )
		self.m_grid2.SetColLabelValue( 2, u"Z" )
		self.m_grid2.SetColLabelValue( 3, u"CA1" )
		self.m_grid2.SetColLabelValue( 4, u"CA2" )
		self.m_grid2.SetColLabelValue( 5, wx.EmptyString )
		self.m_grid2.SetColLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )
        
		ncol=1050
		for i in range(ncol):
        		self.m_grid2.SetColLabelValue( i+5, str(i+1) )
  
		# Rows
		self.m_grid2.EnableDragRowSize( True )
		self.m_grid2.SetRowLabelSize( 40 )
		self.m_grid2.SetRowLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Label Appearance
		self.m_grid2.SetLabelFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )

		# Cell Defaults
		self.m_grid2.SetDefaultCellFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )
		self.m_grid2.SetDefaultCellAlignment( wx.ALIGN_LEFT, wx.ALIGN_TOP )
		self.m_grid2.SetFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )

		sbSizer16.Add( self.m_grid2, 0, wx.ALL, 5 )

		self.m_button6 = wx.Button( sbSizer16.GetStaticBox(), wx.ID_ANY, u"Get Distances", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		sbSizer16.Add( self.m_button6, 0, wx.ALIGN_CENTER|wx.ALL, 5 )


		gbSizer2.Add( sbSizer16, wx.GBPosition( 0, 1 ), wx.GBSpan( 2, 1 ), wx.EXPAND|wx.TOP, 5 )

		sbSizer18 = wx.StaticBoxSizer( wx.StaticBox( self, wx.ID_ANY, u"Scores" ), wx.VERTICAL )

		fgSizer2 = wx.FlexGridSizer( 1, 2, 0, 0 )
		fgSizer2.SetFlexibleDirection( wx.BOTH )
		fgSizer2.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )

		self.m_grid21 = wx.grid.Grid( sbSizer18.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition, wx.Size( 368,230 ), 0 )

		# Grid
		self.m_grid21.CreateGrid( maxccs, 7 )
		self.m_grid21.EnableEditing( True )
		self.m_grid21.EnableGridLines( False )
		self.m_grid21.EnableDragGridSize( False )
		self.m_grid21.SetMargins( 0, 0 )

		# Columns
		self.m_grid21.SetColSize( 0, 33 )
		self.m_grid21.SetColSize( 1, 33 )
		self.m_grid21.SetColSize( 2, 55 )
		self.m_grid21.SetColSize( 3, 55 )
		self.m_grid21.SetColSize( 4, 55 )
		self.m_grid21.SetColSize( 5, 55 )
		self.m_grid21.SetColSize( 6, 55 )
		self.m_grid21.EnableDragColMove( False )
		self.m_grid21.EnableDragColSize( True )
		self.m_grid21.SetColLabelSize( 25 )
		self.m_grid21.SetColLabelValue( 0, u"CA1" )
		self.m_grid21.SetColLabelValue( 1, u"CA2" )
		self.m_grid21.SetColLabelValue( 2, u"ave" )
		self.m_grid21.SetColLabelValue( 3, u"stdev" )
		self.m_grid21.SetColLabelValue( 4, u"score" )
		self.m_grid21.SetColLabelValue( 5, u"max" )
		self.m_grid21.SetColLabelValue( 6, u"min" )
		self.m_grid21.SetColLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Rows
		self.m_grid21.EnableDragRowSize( True )
		self.m_grid21.SetRowLabelSize( 40 )
		self.m_grid21.SetRowLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Label Appearance
		self.m_grid21.SetLabelFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )

		# Cell Defaults
		self.m_grid21.SetDefaultCellFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )
		self.m_grid21.SetDefaultCellAlignment( wx.ALIGN_LEFT, wx.ALIGN_TOP )
		self.m_grid21.SetFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )

		fgSizer2.Add( self.m_grid21, 0, wx.ALIGN_CENTER|wx.ALL, 5 )

		self.m_grid3 = wx.grid.Grid( sbSizer18.GetStaticBox(), wx.ID_ANY, wx.DefaultPosition, wx.Size( 180,230 ), 0 )

		# Grid
		self.m_grid3.CreateGrid( maxccs, 3 )
		self.m_grid3.EnableEditing( True )
		self.m_grid3.EnableGridLines( False )
		self.m_grid3.EnableDragGridSize( False )
		self.m_grid3.SetMargins( 0, 0 )

		# Columns
		self.m_grid3.SetColSize( 0, 33 )
		self.m_grid3.SetColSize( 1, 33 )
		self.m_grid3.SetColSize( 2, 55 )
		self.m_grid3.EnableDragColMove( False )
		self.m_grid3.EnableDragColSize( True )
		self.m_grid3.SetColLabelSize( 25 )
		self.m_grid3.SetColLabelValue( 0, u"CA1" )
		self.m_grid3.SetColLabelValue( 1, u"CA2" )
		self.m_grid3.SetColLabelValue( 2, u"score" )
		self.m_grid3.SetColLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Rows
		self.m_grid3.EnableDragRowSize( True )
		self.m_grid3.SetRowLabelSize( 40 )
		self.m_grid3.SetRowLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Label Appearance
		self.m_grid3.SetLabelFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )

		# Cell Defaults
		self.m_grid3.SetDefaultCellFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )
		self.m_grid3.SetDefaultCellAlignment( wx.ALIGN_LEFT, wx.ALIGN_TOP )
		self.m_grid3.SetFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )

		fgSizer2.Add( self.m_grid3, 0, wx.ALIGN_CENTER|wx.ALL, 5 )


		sbSizer18.Add( fgSizer2, 1, 0, 5 )

		bSizer6 = wx.BoxSizer( wx.HORIZONTAL )


		bSizer6.Add( ( 130, 0), 0, 0, 5 )

		self.m_button8 = wx.Button( sbSizer18.GetStaticBox(), wx.ID_ANY, u"Get Scores", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		bSizer6.Add( self.m_button8, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 5 )


		bSizer6.Add( ( 145, 0), 0, 0, 5 )

		self.m_button7 = wx.Button( sbSizer18.GetStaticBox(), wx.ID_ANY, u"Plot", wx.DefaultPosition, wx.Size( 90,30 ), 0 )
		bSizer6.Add( self.m_button7, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 5 )

		self.m_button11 = wx.Button( sbSizer18.GetStaticBox(), wx.ID_ANY, u"Log Plot", wx.DefaultPosition, wx.Size( 90,30 ), 0 )
		bSizer6.Add( self.m_button11, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 5 )


		sbSizer18.Add( bSizer6, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALIGN_CENTER_VERTICAL|wx.EXPAND, 5 )


		gbSizer2.Add( sbSizer18, wx.GBPosition( 1, 0 ), wx.GBSpan( 1, 1 ), wx.EXPAND|wx.LEFT|wx.RIGHT|wx.TOP, 5 )


		self.SetSizer( gbSizer2 )
		self.Layout()
		self.m_menubar1 = wx.MenuBar( 0 )
		self.m_menu1 = wx.Menu()
		self.m_menuItem1 = wx.Menu()
		self.m_menuItem15 = wx.MenuItem( self.m_menuItem1, wx.ID_ANY, u"Distance", wx.EmptyString, wx.ITEM_NORMAL )
		self.m_menuItem1.Append( self.m_menuItem15 )

		self.m_menuItem18 = wx.MenuItem( self.m_menuItem1, wx.ID_ANY, u"Score", wx.EmptyString, wx.ITEM_NORMAL )
		self.m_menuItem1.Append( self.m_menuItem18 )

		self.m_menu1.AppendSubMenu( self.m_menuItem1, u"Open" )

		self.m_menu5 = wx.Menu()
		self.m_menuItem17 = wx.MenuItem( self.m_menu5, wx.ID_ANY, u"Distance", wx.EmptyString, wx.ITEM_NORMAL )
		self.m_menu5.Append( self.m_menuItem17 )

		self.m_menuItem171 = wx.MenuItem( self.m_menu5, wx.ID_ANY, u"Score", wx.EmptyString, wx.ITEM_NORMAL )
		self.m_menu5.Append( self.m_menuItem171 )

		self.m_menu1.AppendSubMenu( self.m_menu5, u"Save" )

		self.m_menuItem4 = wx.MenuItem( self.m_menu1, wx.ID_ANY, u"Exit", wx.EmptyString, wx.ITEM_NORMAL )
		self.m_menu1.Append( self.m_menuItem4 )

		self.m_menubar1.Append( self.m_menu1, u"File" )

		self.m_menu3 = wx.Menu()
		self.m_menuItem201 = wx.MenuItem( self.m_menu3, wx.ID_ANY, u"Progress plot", wx.EmptyString, wx.ITEM_NORMAL )
		self.m_menu3.Append( self.m_menuItem201 )

		self.m_menuItem7 = wx.MenuItem( self.m_menu3, wx.ID_ANY, u"Cis peptides", wx.EmptyString, wx.ITEM_NORMAL )
		self.m_menu3.Append( self.m_menuItem7 )
        
		self.m_menubar1.Append( self.m_menu3, u"Utilities" )

		self.SetMenuBar( self.m_menubar1 )

		self.m_statusBar1 = self.CreateStatusBar( 1, wx.STB_SIZEGRIP, wx.ID_ANY )

		self.Centre( wx.BOTH )

		# Connect Events
		self.m_button1.Bind( wx.EVT_BUTTON, self.OnOpenPDB )
		self.m_button16.Bind( wx.EVT_BUTTON, self.OnOpenCIF )
		self.m_button81.Bind( wx.EVT_BUTTON, self.OnClear )
		self.m_button6.Bind( wx.EVT_BUTTON, self.OnDistCalc )
		self.m_button8.Bind( wx.EVT_BUTTON, self.OnScoreCalc )
		self.m_button7.Bind( wx.EVT_BUTTON, self.OnExpand )
		self.m_button11.Bind( wx.EVT_BUTTON, self.OnExpand2 )
		self.Bind( wx.EVT_MENU, self.OnLoadDist, id = self.m_menuItem15.GetId() )
		self.Bind( wx.EVT_MENU, self.OnLoadScore, id = self.m_menuItem18.GetId() )
		self.Bind( wx.EVT_MENU, self.OnSaveDist2, id = self.m_menuItem17.GetId() )
		self.Bind( wx.EVT_MENU, self.OnSaveScore2, id = self.m_menuItem171.GetId() )
		self.Bind( wx.EVT_MENU, self.OnExit, id = self.m_menuItem4.GetId() )
		self.Bind( wx.EVT_MENU, self.OnAverageScore, id = self.m_menuItem201.GetId() )
		self.Bind( wx.EVT_MENU, self.OnCisPeptides, id = self.m_menuItem7.GetId() )


#	def __del__( self ):
#		pass
	
	
	def OnLoadDist( self, event ):
		self.SetStatusText("")
		self.m_richText2.Clear()
		self.m_textCtrl31.Clear()
		self.m_textCtrl6.Clear()
		self.m_textCtrl2.Clear()
		self.m_textCtrl7.Clear()
		self.m_textCtrl3.Clear()
		self.m_textCtrl4.Clear()
		self.m_grid2.ClearGrid()
		self.m_grid21.ClearGrid()
		self.m_grid3.ClearGrid()
#		for id in range(0, 50):
#		    self.m_checkList1.Check(id, False)
      
		self.dirName = ''
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    xlsxlrd = xlrd.open_workbook(os.path.join(self.dirName, self.fileName))
		    sheet_1 = xlsxlrd.sheet_by_index(0)
		    self.m_textCtrl3.SetValue(self.fileName)
		    self.SetStatusText("Loading ...")
		    for i in range(sheet_1.nrows):
                      ca1=int(sheet_1.cell(i,0).value)
                      self.m_grid2.SetCellValue(i, 3, str(ca1))
                      self.m_grid21.SetCellValue(i, 0, str(ca1))
                      self.m_grid3.SetCellValue(i, 0, str(ca1))
                      ca2=int(sheet_1.cell(i,1).value)
                      self.m_grid2.SetCellValue(i, 4, str(ca2))
                      self.m_grid21.SetCellValue(i, 1, str(ca2))
                      self.m_grid3.SetCellValue(i, 1, str(ca2))
		    for col in range(2,sheet_1.ncols):
                      for row in range(sheet_1.nrows):
                          cc=str(sheet_1.cell(row,col).value)
                          self.m_grid2.SetCellValue(row, col+3, cc[0:11])
                      
		    colnum = sheet_1.ncols-2
		    self.m_textCtrl31.SetValue(str(colnum))

		    rownum = sheet_1.nrows
		    self.m_textCtrl2.SetValue(str(rownum))

		    frame.SetStatusText("Done")

	def OnLoadScore( self, event ):
		self.m_richText2.Clear()
		self.m_textCtrl31.Clear()
		self.m_textCtrl6.Clear()
		self.m_textCtrl2.Clear()
		self.m_textCtrl7.Clear()
		self.m_textCtrl3.Clear()
		self.m_textCtrl4.Clear()
		self.m_grid2.ClearGrid()
		self.m_grid21.ClearGrid()
		self.m_grid3.ClearGrid()
                      
		self.dirName = ''
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    xlsxlrd = xlrd.open_workbook(os.path.join(self.dirName, self.fileName))
		    sheet_1 = xlsxlrd.sheet_by_index(0)
		    self.m_textCtrl3.SetValue(self.fileName)
		    frame.SetStatusText("Loading ...")
		    for i in range(sheet_1.nrows):
                      ca1=int(sheet_1.cell(i,0).value)
                      self.m_grid21.SetCellValue(i, 0, str(ca1))
                      self.m_grid3.SetCellValue(i, 0, str(ca1))
                      ca2=int(sheet_1.cell(i,1).value)
                      self.m_grid21.SetCellValue(i, 1, str(ca2))
                      self.m_grid3.SetCellValue(i, 1, str(ca2))
		    for col in range(2,sheet_1.ncols):
                      for row in range(sheet_1.nrows):
                          cc=str(sheet_1.cell(row,col).value)
                          self.m_grid21.SetCellValue(row, col, cc[0:11])
                          if col==4:
                              self.m_grid3.SetCellValue(row, 2, cc[0:11])
                              
		    rownum = sheet_1.nrows
		    self.m_textCtrl2.SetValue(str(rownum))

		    frame.SetStatusText("Done")
      
	
	def OnSaveDist2( self, event ):
		self.SetStatusText("")
		self.SetStatusText("Saving ...")
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    wb = Workbook()
#		    ws = wb.create_sheet(0)
		    ws = wb.active
		    rownum=int(self.m_textCtrl2.GetValue())
		    colnum = int(self.m_textCtrl31.GetValue())
		    for i in range(0, rownum):
                      ca1 = int(self.m_grid2.GetCellValue(i, 3))
                      ca2 = int(self.m_grid2.GetCellValue(i, 4))
                      ws.cell(column=1, row=i+1).value=ca1
                      ws.cell(column=2, row=i+1).value=ca2
                      for j in range(5, colnum+5):
                          dist = float(self.m_grid2.GetCellValue(i, j))
                          ws.cell(column=j-2, row=i+1).value=dist
		    wb.save(os.path.join(self.dirName, self.fileName))
		dialog.Destroy()
		self.SetStatusText("Done")


	def OnSaveScore2( self, event ):
		self.SetStatusText("")
		self.SetStatusText("Saving ...")
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    wb = Workbook()
		    ws = wb.active
		    rownum=int(self.m_textCtrl2.GetValue())
		    for i in range(0, rownum):
                      ca1 = float(self.m_grid21.GetCellValue(i, 0))
                      ca2 = float(self.m_grid21.GetCellValue(i, 1))
                      ave = float(self.m_grid21.GetCellValue(i, 2))
                      dev = float(self.m_grid21.GetCellValue(i, 3))
                      score = float(self.m_grid21.GetCellValue(i, 4))
                      mx = float(self.m_grid21.GetCellValue(i, 5))
                      mn = float(self.m_grid21.GetCellValue(i, 6))
                      ws.cell(column=1, row=i+1).value=ca1
                      ws.cell(column=2, row=i+1).value=ca2
                      ws.cell(column=3, row=i+1).value=ave
                      ws.cell(column=4, row=i+1).value=dev
                      ws.cell(column=5, row=i+1).value=score
                      ws.cell(column=6, row=i+1).value=mx
                      ws.cell(column=7, row=i+1).value=mn
		    wb.save(os.path.join(self.dirName, self.fileName))
		dialog.Destroy()
		self.SetStatusText("Done")



	def OnExit( self, event ):
		dialog = wx.MessageDialog( self, "exit?", "Exit Messege", wx.OK)
		dialog.ShowModal()
		dialog.Destroy()
		self.Close()
#		exit()



	def OnAverageScore( self, event ):
		cutofflow = float(panel2.m_textCtrl19.GetValue())
		dcutofflow = float(panel2.m_textCtrl18.GetValue())
		cutoff = float(panel2.m_textCtrl16.GetValue())
		dcutoff = float(panel2.m_textCtrl15.GetValue())
		distlow = float(panel2.m_textCtrl12.GetValue())
		disthigh = float(panel2.m_textCtrl11.GetValue())
		scorelow = float(panel2.m_textCtrl14.GetValue())
		scorehigh = float(panel2.m_textCtrl13.GetValue())  
#		scorelim = float(panel2.m_textCtrl17.GetValue())
		colnum = int(self.m_textCtrl31.GetValue())
		rownum = int(self.m_textCtrl2.GetValue())
		panel1 = Graph_frame(self) 
		panel1.Show()

		self.figure2 = Figure((8.4, 6.8), dpi=100)
		self.axes = self.figure2.add_subplot(111)
		self.axes.set_xlim(dcutofflow,dcutoff)
		self.axes.set_ylim(cutofflow,cutoff)
		self.axes.set_position(pos = [0.15,0.11,0.8,0.84])
		self.axes.set_xlabel('Number of Chains', fontsize=20)
		self.axes.set_ylabel('Average Score', fontsize=20)
		self.axes.tick_params(axis='both', labelsize=16)

		panel8 = Data_frame2(self)
		panel8.Show()                  
		panel8.SetStatusText("Calculating ...")
		for j in range(3, colnum+1):
		    y = 0
		    i = 0
		    for i in range(0, rownum):
                      s = 0
                      n = 1
                      for col in range(5, j+5):
                              d = float(self.m_grid2.GetCellValue(i, col))
                              s = s + d
                              ave = s/n
                              avecell = str(ave)
                              n = n + 1                              
                      self.m_grid21.SetCellValue(i, 2, avecell[0:11])
                      

		    for i in range(0, rownum):
                      sasqsum = 0
                      n = 1
                      ave = float(self.m_grid21.GetCellValue(i, 2))
                      for col in range(5, j+5):
                              d = float(self.m_grid2.GetCellValue(i, col))
                              sasq = (d - ave)*(d - ave)
                              sasqsum = sasqsum + sasq
                              b = sasqsum/float(n) 
                              dev = math.sqrt(b)
                              devcell = str(Decimal(dev))
                              n = n + 1
                      self.m_grid21.SetCellValue(i, 3, devcell[0:11])
                      

		    for i in range(0, rownum):          
                      ave = float(self.m_grid21.GetCellValue(i, 2))
                      dev = float(self.m_grid21.GetCellValue(i, 3))
                      if dev == 0.0:
                          dev = 0.00001
                      score = ave/dev
                      scorecell =str(score)
                      self.m_grid21.SetCellValue(i, 4, scorecell[0:11])
                      self.m_grid3.SetCellValue(i, 2, scorecell[0:11])
                              
                      y = y + score

		    avescore = y / rownum
		    x = j
		    y = avescore      
		    if x > distlow:
                                  if x < disthigh:
                                      if y > scorelow:
                                          if y < scorehigh:
                                              self.axes.plot(x,y,"ro-",markersize=2)

		    panel8.m_grid6.SetCellValue(j-1, 0, str(j))
		    panel8.m_grid6.SetCellValue(j-1, 1, str(y))

		self.canvas = FigCanvas(panel1.m_panel3, wx.ID_ANY, self.figure2)
		panel8.SetStatusText("Done")

	def OnCisPeptides( self, event ):

		rownum = int(self.m_textCtrl2.GetValue())

		panel81 = Data_frame21(self)
		panel81.Show()                  
		panel81.SetStatusText("Calculating ...")
		cis = 0
        
		for i in range(0, rownum):
                      adj = float(self.m_grid21.GetCellValue(i, 6))
                      if adj < 3.3:
                          ave = float(self.m_grid21.GetCellValue(i, 2))
                          ave = str(ave)
                          panel81.m_grid61.SetCellValue(cis, 2, ave[0:11])
                          ca1 = int(self.m_grid21.GetCellValue(i, 0))
                          ca2 = int(self.m_grid21.GetCellValue(i, 1))
                          score = float(self.m_grid21.GetCellValue(i, 4))
                          panel81.m_grid61.SetCellValue(cis, 0, str(ca1))
                          panel81.m_grid61.SetCellValue(cis, 1, str(ca2))
                          score = str(score)
                          panel81.m_grid61.SetCellValue(cis, 3, score[0:11])
                          cis = cis+1
 
		panel81.SetStatusText("Done")

	def OnClear( self, event ):
		self.m_richText2.Clear()
		self.m_textCtrl31.Clear()
		self.m_textCtrl6.Clear()
		self.m_textCtrl2.Clear()
		self.m_textCtrl7.Clear()
		self.m_textCtrl3.Clear()
		self.m_textCtrl4.Clear()
		self.m_grid2.ClearGrid()
		self.m_grid21.ClearGrid()
		self.m_grid3.ClearGrid()

#		self.figure = Figure((3.2, 2.8), dpi=100)
#		self.canvas = FigCanvas(self.m_panel1, wx.ID_ANY, self.figure)



	def OnScoreCalc(self, event): 
		    self.SetStatusText("")
		    self.SetStatusText("Calculating ...")
		    i = 0
		    rownum = int(self.m_textCtrl2.GetValue())
		    colnum = int(self.m_textCtrl31.GetValue())

#		    for id in range(0, colnum):
#                      self.m_checkList1.Check(id)
                      
		    for i in range(0, rownum):
                      s = 0
                      n = 1
                      for col in range(5, colnum+5):
#                          if self.m_checkList1.IsChecked(col-5):
                              d = float(self.m_grid2.GetCellValue(i, col))
                              s = s + d
                              ave = s/n
                              avecell = str(ave)
                              n = n + 1                              
                      self.m_grid21.SetCellValue(i, 2, avecell[0:11])
                      

		    for i in range(0, rownum):
                      sasqsum = 0
                      n = 1
                      ave = float(self.m_grid21.GetCellValue(i, 2))
                      for col in range(5, colnum+5):
#                          if self.m_checkList1.IsChecked(col-5):
                              d = float(self.m_grid2.GetCellValue(i, col))
                              sasq = (d - ave)*(d - ave)
                              sasqsum = sasqsum + sasq
                              b = sasqsum/float(n) 
                              dev = math.sqrt(b)
                              devcell = str(Decimal(dev))
                              n = n + 1
                      self.m_grid21.SetCellValue(i, 3, devcell[0:11])
                      

		    for i in range(0, rownum):          
                      ave = float(self.m_grid21.GetCellValue(i, 2))
                      dev = float(self.m_grid21.GetCellValue(i, 3))
                      if dev == 0.0:
                          dev = 0.00001
                      score = ave/dev
                      scorecell =str(score)
                      self.m_grid21.SetCellValue(i, 4, scorecell[0:11])
                      self.m_grid3.SetCellValue(i, 2, scorecell[0:11])
		    
		    for i in range(0, rownum):
                      max = 0          
                      for col in range(5, colnum+5):
#                          if self.m_checkList1.IsChecked(col-5):
                              d = float(self.m_grid2.GetCellValue(i, col))
                              if max < d:
                                  max = d                    
                              maxcell = str(max)
                      self.m_grid21.SetCellValue(i, 5, maxcell[0:11])

		    for i in range(0, rownum):
                      min = 200          
                      for col in range(5, colnum+5):
#                          if self.m_checkList1.IsChecked(col-5):
                              d = float(self.m_grid2.GetCellValue(i, col))
                              if min > d:
                                  min = d                    
                              mincell = str(min)
                      self.m_grid21.SetCellValue(i, 6, mincell[0:11])
		    self.SetStatusText("Done")

	def OnOpenPDB( self, event ):
		self.m_richText2.Clear()
		self.dirName = ''
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName1 = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    file1 = open(os.path.join(self.dirName, self.fileName1), 'r')
		    self.m_textCtrl3.SetValue(self.fileName1)
		    rownum=0
		chainid = str(self.m_textCtrl4.GetValue())
		if chainid == "":
		    for line in file1:
                      head=line[0:6]
                      if head.find('ATOM') != -1:
                          atomname=line[12:16]
                          if atomname.find('CA') != -1:
                              alt=line[16:17]
                              if alt.find('B') != -1:
                                  print(str(line[17:29]))
                              elif alt.find('C') != -1:
                                  print(str(line[17:29]))
                              elif alt.find('D') != -1:
                                  print(str(line[17:29]))
                              else:
                                  self.m_grid2.SetCellValue(rownum, 0, str(line[30:38]))
                                  self.m_grid2.SetCellValue(rownum, 1, str(line[38:46]))
                                  self.m_grid2.SetCellValue(rownum, 2, str(line[46:54]))
                                  self.m_richText2.AppendText(str(line))
                                  rownum=rownum+1
                      if head.find('HETATM') != -1:
                          atomname=line[12:16]
                          if atomname.find('CA') != -1:
                              alt=line[16:17]
                              if alt.find('B') != -1:
                                  print(str(line[17:29]))
                              elif alt.find('C') != -1:
                                  print(str(line[17:29]))
                              else:
                                  self.m_grid2.SetCellValue(rownum, 0, str(line[30:38]))
                                  self.m_grid2.SetCellValue(rownum, 1, str(line[38:46]))
                                  self.m_grid2.SetCellValue(rownum, 2, str(line[46:54]))
                                  self.m_richText2.AppendText(str(line))
                                  rownum=rownum+1                           
		    resnum=rownum
		    inires=1
		    self.m_textCtrl6.SetValue(str(resnum))
		    ccnum=int(resnum*(resnum-1)*0.5)
		    self.m_textCtrl2.SetValue(str(ccnum))
		    self.m_textCtrl7.SetValue(str(inires))
                       

		else:
		    for line in file1:
                      head=line[0:6]
                      chain=str(line[21:23])
                      if chain.find(chainid) != -1:
                          if head.find('ATOM') != -1:
                              atomname=line[12:16]
                              if atomname.find('CA') != -1:
                                  alt=line[16:17]
                                  if alt.find('B') != -1:
                                      print(str(line[17:29]))
                                  elif alt.find('C') != -1:
                                      print(str(line[17:29]))
                                  elif alt.find('D') != -1:
                                      print(str(line[17:29]))
                                  else:
                                      self.m_grid2.SetCellValue(rownum, 0, str(line[30:38]))
                                      self.m_grid2.SetCellValue(rownum, 1, str(line[38:46]))
                                      self.m_grid2.SetCellValue(rownum, 2, str(line[46:54]))
                                      self.m_richText2.AppendText(str(line))
                                      rownum=rownum+1
                          if head.find('HETATM') != -1:
                              atomname=line[12:16]
                              if atomname.find('CA') != -1:
                                  alt=line[16:17]
                                  if alt.find('B') != -1:
                                      print(str(line[17:29]))
                                  elif alt.find('C') != -1:
                                      print(str(line[17:29]))
                                  else:
                                      self.m_grid2.SetCellValue(rownum, 0, str(line[30:38]))
                                      self.m_grid2.SetCellValue(rownum, 1, str(line[38:46]))
                                      self.m_grid2.SetCellValue(rownum, 2, str(line[46:54]))
                                      self.m_richText2.AppendText(str(line))
                                      rownum=rownum+1                           
		    resnum=rownum
		    self.m_textCtrl6.SetValue(str(resnum))
		    ccnum=int(resnum*(resnum-1)*0.5)
		    self.m_textCtrl2.SetValue(str(ccnum))
		    inires=1
		    self.m_textCtrl7.SetValue(str(inires))

	def OnOpenCIF( self, event ):
		self.m_richText2.Clear()
		self.dirName = ''
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName1 = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    file1 = open(os.path.join(self.dirName, self.fileName1), 'r')
		    self.m_textCtrl3.SetValue(self.fileName1)
		    rownum=0
		chainid = str(self.m_textCtrl4.GetValue())
		if chainid == "":
		    for line in file1:
                      head=line[0:6]
                      if head.find('ATOM') != -1:
                          atomname=line[14:18]
                          if atomname.find('CA') != -1:
                              alt=line[18:19]
                              if alt.find('B') != -1:
                                  print(str(line[20:32]))
                              elif alt.find('C') != -1:
                                  print(str(line[20:32]))
                              elif alt.find('D') != -1:
                                  print(str(line[20:32]))
                              else:
                                  self.m_grid2.SetCellValue(rownum, 0, str(line[34:43]))
                                  self.m_grid2.SetCellValue(rownum, 1, str(line[43:52]))
                                  self.m_grid2.SetCellValue(rownum, 2, str(line[52:61]))
                                  self.m_richText2.AppendText(str(line))
                                  rownum=rownum+1
                      if head.find('HETATM') != -1:
                          atomname=line[14:18]
                          if atomname.find('CA') != -1:
                              alt=line[18:19]
                              if alt.find('B') != -1:
                                  print(str(line[20:32]))
                              elif alt.find('C') != -1:
                                  print(str(line[20:32]))
                              else:
                                  self.m_grid2.SetCellValue(rownum, 0, str(line[34:43]))
                                  self.m_grid2.SetCellValue(rownum, 1, str(line[43:52]))
                                  self.m_grid2.SetCellValue(rownum, 2, str(line[52:61]))
                                  self.m_richText2.AppendText(str(line))
                                  rownum=rownum+1                           
		    resnum=rownum
		    inires=1
		    self.m_textCtrl6.SetValue(str(resnum))
		    ccnum=int(resnum*(resnum-1)*0.5)
		    self.m_textCtrl2.SetValue(str(ccnum))
		    self.m_textCtrl7.SetValue(str(inires))
                       

		else:
		    for line in file1:
                      head=line[0:6]
                      chain=str(line[25:27])
                      if chain.find(chainid) != -1:
                          if head.find('ATOM') != -1:
                              atomname=line[14:18]
                              if atomname.find('CA') != -1:
                                  alt=line[18:19]
                                  if alt.find('B') != -1:
                                      print(str(line[20:32]))
                                  elif alt.find('C') != -1:
                                      print(str(line[20:32]))
                                  elif alt.find('D') != -1:
                                      print(str(line[20:32]))
                                  else:
                                      self.m_grid2.SetCellValue(rownum, 0, str(line[34:43]))
                                      self.m_grid2.SetCellValue(rownum, 1, str(line[43:52]))
                                      self.m_grid2.SetCellValue(rownum, 2, str(line[52:61]))
                                      self.m_richText2.AppendText(str(line))
                                      rownum=rownum+1
                          if head.find('HETATM') != -1:
                              atomname=line[12:16]
                              if atomname.find('CA') != -1:
                                  alt=line[16:17]
                                  if alt.find('B') != -1:
                                      print(str(line[20:32]))
                                  elif alt.find('C') != -1:
                                      print(str(line[20:32]))
                                  else:
                                      self.m_grid2.SetCellValue(rownum, 0, str(line[34:43]))
                                      self.m_grid2.SetCellValue(rownum, 1, str(line[43:52]))
                                      self.m_grid2.SetCellValue(rownum, 2, str(line[52:61]))
                                      self.m_richText2.AppendText(str(line))
                                      rownum=rownum+1                           
		    resnum=rownum
		    self.m_textCtrl6.SetValue(str(resnum))
		    ccnum=int(resnum*(resnum-1)*0.5)
		    self.m_textCtrl2.SetValue(str(ccnum))
		    inires=1
		    self.m_textCtrl7.SetValue(str(inires))
	def OnDistCalc( self, event ):
		    maxccs = 100000
		    maxchains = 1050
		    logger = logging.getLogger()
		    fhandler = logging.FileHandler(filename='dsa.log', mode='a')
		    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
		    fhandler.setFormatter(formatter)
		    logger.addHandler(fhandler)
		    logger.setLevel(logging.DEBUG)
		    resnum=int(self.m_textCtrl6.GetValue())
		    dist = 0
		    posi = -1            
		    ccposi=5       
		    inires=int(self.m_textCtrl7.GetValue())
		    for isfile in range(5, maxchains+4):
                      judge2=self.m_grid2.GetCellValue(0, isfile)
                      if judge2 != "":
                          ccposi=isfile+1 
		    for ca1 in range(inires, inires+resnum-1):
                      x1 = float(self.m_grid2.GetCellValue(ca1-1, 0))
                      y1 = float(self.m_grid2.GetCellValue(ca1-1, 1))
                      z1 = float(self.m_grid2.GetCellValue(ca1-1, 2))
                      for ca2 in range(ca1+1, inires+resnum):
                          posi = posi+1
                          if posi < maxccs:
                              self.m_grid2.SetCellValue(posi, 3, str(ca1))
                              self.m_grid2.SetCellValue(posi, 4, str(ca2))
                              self.m_grid21.SetCellValue(posi, 0, str(ca1))
                              self.m_grid21.SetCellValue(posi, 1, str(ca2))
                              self.m_grid3.SetCellValue(posi, 0, str(ca1))
                              self.m_grid3.SetCellValue(posi, 1, str(ca2))
                              x2 = float(self.m_grid2.GetCellValue(ca2-1, 0))
                              y2 = float(self.m_grid2.GetCellValue(ca2-1, 1))
                              z2 = float(self.m_grid2.GetCellValue(ca2-1, 2))
                              dist = float(((x1-x2)**2+(y1-y2)**2+(z1-z2)**2)**0.5)
                              dist1 = str(dist)
                              self.m_grid2.SetCellValue(posi, ccposi, dist1[0:11])
		    filenum=ccposi-4
		    self.m_textCtrl31.SetValue(str(filenum))

#		    for id in range(0, filenum):
#                      self.m_checkList1.Check(id)

		    logging.info(str(filenum))
		    logging.info(int(self.m_textCtrl7.GetValue()))
		    logging.info(self.m_textCtrl3.GetValue())	

     

	def OnExpand(self, event): 


		    cutofflow = float(panel2.m_textCtrl19.GetValue())
		    dcutofflow = float(panel2.m_textCtrl18.GetValue())
		    cutoff = float(panel2.m_textCtrl16.GetValue())
		    dcutoff = float(panel2.m_textCtrl15.GetValue())
		    distlow = float(panel2.m_textCtrl12.GetValue())
		    disthigh = float(panel2.m_textCtrl11.GetValue())
		    scorelow = float(panel2.m_textCtrl14.GetValue())
		    scorehigh = float(panel2.m_textCtrl13.GetValue())  

		    panel1 = Graph_frame(self) 
		    panel1.Show()
		    rownum = int(self.m_textCtrl2.GetValue())

		    self.figure2 = Figure((8.4, 6.8), dpi=100)
		    self.axes = self.figure2.add_subplot(111)
		    self.axes.set_xlim(dcutofflow,dcutoff)
		    self.axes.set_ylim(cutofflow,cutoff)
		    self.axes.set_position(pos = [0.15,0.11,0.8,0.84])
		    self.axes.set_xlabel('Distance (\u00c5)', fontsize=20)
		    self.axes.set_ylabel('Score', fontsize=20)
		    self.axes.tick_params(axis='both', labelsize=16)
#                  self.axes.set_yticklabels(size=18)
                  
		    for i in range(0, rownum):                
                      x = float(self.m_grid21.GetCellValue(i, 2))
                      y = float(self.m_grid21.GetCellValue(i, 4))
                      if x > distlow:
                                  if x < disthigh:
                                      if y > scorelow:
                                          if y < scorehigh:
                                              self.axes.plot(x,y,"ro-",markersize=2)
		    self.canvas = FigCanvas(panel1.m_panel3, wx.ID_ANY, self.figure2)      
	

	def OnExpand2(self, event): 


		    cutofflow = float(panel2.m_textCtrl19.GetValue())
		    dcutofflow = float(panel2.m_textCtrl18.GetValue())
		    cutoff = float(panel2.m_textCtrl16.GetValue())
		    dcutoff = float(panel2.m_textCtrl15.GetValue())
		    distlow = float(panel2.m_textCtrl12.GetValue())
		    disthigh = float(panel2.m_textCtrl11.GetValue())
		    scorelow = float(panel2.m_textCtrl14.GetValue())
		    scorehigh = float(panel2.m_textCtrl13.GetValue())  

		    panel1 = Graph_frame(self) 
		    panel1.Show()
		    rownum = int(self.m_textCtrl2.GetValue())

		    self.figure2 = Figure((8.4, 6.8), dpi=100)
		    self.axes = self.figure2.add_subplot(111)
		    self.axes.set_xlim(dcutofflow,dcutoff)
		    self.axes.set_ylim(cutofflow,cutoff)
		    self.axes.set_position(pos = [0.15,0.11,0.8,0.84])
		    self.axes.set_xlabel('Distance (\u00c5)', fontsize=20)
		    self.axes.set_ylabel('Score', fontsize=20)
		    self.axes.set_yscale("log")
		    self.axes.tick_params(axis='both', labelsize=16)
#                  self.axes.set_yticklabels(size=18)
                  
		    for i in range(0, rownum):                
                      x = float(self.m_grid21.GetCellValue(i, 2))
                      y = float(self.m_grid21.GetCellValue(i, 4))
                      if x > distlow:
                                  if x < disthigh:
                                      if y > scorelow:
                                          if y < scorehigh:
                                              self.axes.plot(x,y,"ro-",markersize=2)
		    self.canvas = FigCanvas(panel1.m_panel3, wx.ID_ANY, self.figure2)      
              

                   
###########################################################################
## Class MyFrame4
###########################################################################

class Graph_frame ( wx.Frame ):

	def __init__( self, parent ):
		wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Plot", pos = wx.DefaultPosition, size = wx.Size( 860,810 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

		self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

		bSizer9 = wx.BoxSizer( wx.VERTICAL )

		self.m_panel3 = wx.Panel( self, wx.ID_ANY, wx.DefaultPosition, wx.Size( -1,830 ), wx.BORDER_SUNKEN|wx.TAB_TRAVERSAL )
		bSizer9.Add( self.m_panel3, 1, wx.EXPAND |wx.ALL, 5 )

		self.m_button9 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( 90,30 ), 0 )
		bSizer9.Add( self.m_button9, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_button12 = wx.Button( self, wx.ID_ANY, u"Close", wx.DefaultPosition, wx.Size( 90,30 ), 0 )
		bSizer9.Add( self.m_button12, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


		self.SetSizer( bSizer9 )
		self.Layout()

		self.Centre( wx.BOTH )

		# Connect Events
		self.m_button9.Bind( wx.EVT_BUTTON, self.OnSaveFig2 )
		self.m_button12.Bind( wx.EVT_BUTTON, self.OnExpandClose )

	def __del__( self ):
		pass
	
	# Virtual event handlers, overide them in your derived class
	def OnExpandClose( self, event ):
		self.Close()

	def OnSaveFig2( self, event ):
		self.dirName = ''
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    frame.figure2.savefig(os.path.join(self.dirName, self.fileName),dpi=300,transparent=True)
		dialog.Destroy()



###########################################################################
## Class MyFrame41
###########################################################################

class Param_frame ( wx.Frame ):

	def __init__( self, parent ):
		wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Parameters", pos = wx.DefaultPosition, size = wx.Size( 250,370 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

		self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

		bSizer8 = wx.BoxSizer( wx.VERTICAL )

		gSizer3 = wx.GridSizer( 0, 2, 0, 0 )

		self.m_staticText23 = wx.StaticText( self, wx.ID_ANY, u"Frame x high", wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTER_HORIZONTAL )
		self.m_staticText23.Wrap( -1 )

		gSizer3.Add( self.m_staticText23, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText10 = wx.StaticText( self, wx.ID_ANY, u"Frame y high", wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTER_HORIZONTAL )
		self.m_staticText10.Wrap( -1 )

		gSizer3.Add( self.m_staticText10, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_textCtrl15 = wx.TextCtrl( self, wx.ID_ANY, u"60", wx.DefaultPosition, wx.Size( 50,-1 ), wx.TE_CENTER )
		gSizer3.Add( self.m_textCtrl15, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_textCtrl16 = wx.TextCtrl( self, wx.ID_ANY, u"700", wx.DefaultPosition, wx.Size( 50,-1 ), wx.TE_CENTER )
		gSizer3.Add( self.m_textCtrl16, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText11 = wx.StaticText( self, wx.ID_ANY, u"Frame x low", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticText11.Wrap( -1 )

		gSizer3.Add( self.m_staticText11, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText12 = wx.StaticText( self, wx.ID_ANY, u"Frame y low", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticText12.Wrap( -1 )

		gSizer3.Add( self.m_staticText12, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_textCtrl18 = wx.TextCtrl( self, wx.ID_ANY, u"0", wx.DefaultPosition, wx.Size( 50,-1 ), wx.TE_CENTER )
		gSizer3.Add( self.m_textCtrl18, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_textCtrl19 = wx.TextCtrl( self, wx.ID_ANY, u"0", wx.DefaultPosition, wx.Size( 50,-1 ), wx.TE_CENTER )
		gSizer3.Add( self.m_textCtrl19, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


		bSizer8.Add( gSizer3, 0, wx.EXPAND, 5 )

		self.m_staticline1 = wx.StaticLine( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.LI_HORIZONTAL )
		bSizer8.Add( self.m_staticline1, 0, wx.EXPAND |wx.ALL, 5 )

		gSizer4 = wx.GridSizer( 0, 2, 0, 0 )

		self.m_staticText13 = wx.StaticText( self, wx.ID_ANY, u"Plot x high", wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTER_HORIZONTAL )
		self.m_staticText13.Wrap( -1 )

		gSizer4.Add( self.m_staticText13, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText14 = wx.StaticText( self, wx.ID_ANY, u"Plot y high", wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTER_HORIZONTAL )
		self.m_staticText14.Wrap( -1 )

		gSizer4.Add( self.m_staticText14, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_textCtrl11 = wx.TextCtrl( self, wx.ID_ANY, u"2000", wx.DefaultPosition, wx.Size( 50,-1 ), wx.TE_CENTER )
		gSizer4.Add( self.m_textCtrl11, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_textCtrl13 = wx.TextCtrl( self, wx.ID_ANY, u"900000", wx.DefaultPosition, wx.Size( 50,-1 ), wx.TE_CENTER )
		gSizer4.Add( self.m_textCtrl13, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText15 = wx.StaticText( self, wx.ID_ANY, u"Plot x low", wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTER_HORIZONTAL )
		self.m_staticText15.Wrap( -1 )

		gSizer4.Add( self.m_staticText15, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_staticText16 = wx.StaticText( self, wx.ID_ANY, u"Plot y low", wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTER_HORIZONTAL )
		self.m_staticText16.Wrap( -1 )

		gSizer4.Add( self.m_staticText16, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_textCtrl12 = wx.TextCtrl( self, wx.ID_ANY, u"0", wx.DefaultPosition, wx.Size( 50,-1 ), wx.TE_CENTER )
		gSizer4.Add( self.m_textCtrl12, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_textCtrl14 = wx.TextCtrl( self, wx.ID_ANY, u"0", wx.DefaultPosition, wx.Size( 50,-1 ), wx.TE_CENTER )
		gSizer4.Add( self.m_textCtrl14, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


		bSizer8.Add( gSizer4, 0, wx.EXPAND, 5 )

		self.m_staticline2 = wx.StaticLine( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.LI_HORIZONTAL )
		bSizer8.Add( self.m_staticline2, 0, wx.EXPAND |wx.ALL, 5 )

		self.m_button12 = wx.Button( self, wx.ID_ANY, u"Close", wx.DefaultPosition, wx.Size( 90,30 ), 0 )
		bSizer8.Add( self.m_button12, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


		self.SetSizer( bSizer8 )
		self.Layout()

		self.Centre( wx.BOTH )

		# Connect Events
		self.m_button12.Bind( wx.EVT_BUTTON, self.OnParamClose )

	def __del__( self ):
		pass
	
	
	def OnParamClose( self, event ):
		self.Close()



###########################################################################
## Class Data_frame2
###########################################################################

class Data_frame2 ( wx.Frame ):

	def __init__( self, parent ):
		maxchains = 1050

		wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Progress data", pos = wx.DefaultPosition, size = wx.Size( 280,750 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

		self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

		bSizer10 = wx.BoxSizer( wx.VERTICAL )

		self.m_grid6 = wx.grid.Grid( self, wx.ID_ANY, wx.DefaultPosition, wx.Size( 260,590 ), 0 )

		# Grid
		self.m_grid6.CreateGrid( maxchains, 2 )
		self.m_grid6.EnableEditing( True )
		self.m_grid6.EnableGridLines( True )
		self.m_grid6.EnableDragGridSize( False )
		self.m_grid6.SetMargins( 0, 0 )

		# Columns
		self.m_grid6.SetColSize( 0, 55 )
		self.m_grid6.SetColSize( 1, 95 )
		self.m_grid6.EnableDragColMove( False )
		self.m_grid6.EnableDragColSize( True )
		self.m_grid6.SetColLabelSize( 30 )
		self.m_grid6.SetColLabelValue( 0, u"chains" )
		self.m_grid6.SetColLabelValue( 1, u"ave.score" )
		self.m_grid6.SetColLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Rows
		self.m_grid6.EnableDragRowSize( True )
		self.m_grid6.SetRowLabelSize( 80 )
		self.m_grid6.SetRowLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Label Appearance
		self.m_grid6.SetLabelFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )

		# Cell Defaults
		self.m_grid6.SetDefaultCellFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )
		self.m_grid6.SetDefaultCellAlignment( wx.ALIGN_LEFT, wx.ALIGN_TOP )
		bSizer10.Add( self.m_grid6, 0, wx.ALL, 5 )

		self.m_button17 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		bSizer10.Add( self.m_button17, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_button18 = wx.Button( self, wx.ID_ANY, u"Close", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		bSizer10.Add( self.m_button18, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


		self.SetSizer( bSizer10 )
		self.Layout()
		self.m_statusBar3 = self.CreateStatusBar( 1, wx.STB_SIZEGRIP, wx.ID_ANY )

		self.Centre( wx.BOTH )

		# Connect Events
		self.m_button17.Bind( wx.EVT_BUTTON, self.OnSavePoints2 )
		self.m_button18.Bind( wx.EVT_BUTTON, self.OnTableClose )

	def __del__( self ):
		pass
	
	
	# Virtual event handlers, overide them in your derived class
	def OnSavePoints2( self, event ):
		maxchains = 1050
		self.dirName = ''
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    book = xlwt.Workbook()
		    newSheet_1 = book.add_sheet("NewSheet_1")
		    rownum=maxchains
		    for i in range(0, rownum):
                      num = str(self.m_grid6.GetCellValue(i, 0))
                      score = str(self.m_grid6.GetCellValue(i, 1))
                      newSheet_1.write(i,0,num)
                      newSheet_1.write(i,1,score)
		    book.save(os.path.join(self.dirName, self.fileName))
		dialog.Destroy()     

	def OnTableClose( self, event ):
		self.Close()	


###########################################################################
## Class Data_frame21
###########################################################################

class Data_frame21 ( wx.Frame ):

	def __init__( self, parent ):
		wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Cis peptide data", pos = wx.DefaultPosition, size = wx.Size( 425,745 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

		self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

		bSizer101 = wx.BoxSizer( wx.VERTICAL )

		self.m_grid61 = wx.grid.Grid( self, wx.ID_ANY, wx.DefaultPosition, wx.Size( 405,590 ), 0 )

		# Grid
		self.m_grid61.CreateGrid( 100, 4 )
		self.m_grid61.EnableEditing( True )
		self.m_grid61.EnableGridLines( True )
		self.m_grid61.EnableDragGridSize( False )
		self.m_grid61.SetMargins( 0, 0 )

		# Columns
		self.m_grid61.SetColSize( 0, 55 )
		self.m_grid61.SetColSize( 1, 55 )
		self.m_grid61.SetColSize( 2, 95 )
		self.m_grid61.SetColSize( 3, 95 )
		self.m_grid61.EnableDragColMove( False )
		self.m_grid61.EnableDragColSize( True )
		self.m_grid61.SetColLabelSize( 30 )
		self.m_grid61.SetColLabelValue( 0, u"CA1" )
		self.m_grid61.SetColLabelValue( 1, u"CA2" )
		self.m_grid61.SetColLabelValue( 2, u"dist" )
		self.m_grid61.SetColLabelValue( 3, u"score" )
		self.m_grid61.SetColLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Rows
		self.m_grid61.EnableDragRowSize( True )
		self.m_grid61.SetRowLabelSize( 80 )
		self.m_grid61.SetRowLabelAlignment( wx.ALIGN_CENTER, wx.ALIGN_CENTER )

		# Label Appearance
		self.m_grid61.SetLabelFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )

		# Cell Defaults
		self.m_grid61.SetDefaultCellFont( wx.Font( 8, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Calibri" ) )
		self.m_grid61.SetDefaultCellAlignment( wx.ALIGN_LEFT, wx.ALIGN_TOP )
		bSizer101.Add( self.m_grid61, 0, wx.ALL, 5 )

		self.m_button171 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		bSizer101.Add( self.m_button171, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )

		self.m_button181 = wx.Button( self, wx.ID_ANY, u"Close", wx.DefaultPosition, wx.Size( -1,30 ), 0 )
		bSizer101.Add( self.m_button181, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


		self.SetSizer( bSizer101 )
		self.Layout()
		self.m_statusBar31 = self.CreateStatusBar( 1, wx.STB_SIZEGRIP, wx.ID_ANY )

		self.Centre( wx.BOTH )

		# Connect Events
		self.m_button171.Bind( wx.EVT_BUTTON, self.OnSavePoints3 )
		self.m_button181.Bind( wx.EVT_BUTTON, self.OnTableClose )

	def __del__( self ):
		pass


	# Virtual event handlers, overide them in your derived class
	def OnSavePoints3( self, event ):
		maxcis = 100
		self.dirName = ''
		dialog = wx.FileDialog(self, "Choose a file", self.dirName, "", "*.*", wx.FD_OPEN)
		if dialog.ShowModal() == wx.ID_OK:
		    self.fileName = dialog.GetFilename()
		    self.dirName = dialog.GetDirectory()
		    book = xlwt.Workbook()
		    newSheet_1 = book.add_sheet("NewSheet_1")
		    rownum=maxcis
		    for i in range(0, rownum):
                      ca1 = str(self.m_grid61.GetCellValue(i, 0))
                      ca2 = str(self.m_grid61.GetCellValue(i, 1))
                      dist = str(self.m_grid61.GetCellValue(i, 2))
                      score = str(self.m_grid61.GetCellValue(i, 3))
                      newSheet_1.write(i,0,ca1)
                      newSheet_1.write(i,1,ca2)
                      newSheet_1.write(i,2,dist)
                      newSheet_1.write(i,3,score)
		    book.save(os.path.join(self.dirName, self.fileName))
		dialog.Destroy()     


	def OnTableClose( self, event ):
		self.Close()	


app = wx.App(False)
frame = MyFrame1(None)
frame.Show()
panel2 = Param_frame(None)
panel2.Show()
#panel8 = Data_frame2(None)

#panel3 = MyFrame5(None)
#import gc
#gc.set_debug(gc.DEBUG_LEAK)
#gc.disable()
#gc.collect()
app.MainLoop()